# coding=utf-8
from __future__ import unicode_literals
import sqlite3
from openpyxl import Workbook

if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    ws.title = "lnkdn"

    conn = sqlite3.connect('linkedin.db')
    c = conn.cursor()

    for i, row in enumerate(c.execute('SELECT * FROM COMPANIES')):
        print row
        ws['A'+str(i+1)] = row[1]
        ws['B'+str(i+1)] = row[2]
        ws['C'+str(i+1)] = row[3]

    conn.close()
    wb.save('lnkdn.xlsx')